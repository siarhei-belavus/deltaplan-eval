#!/usr/bin/env python3
"""
Extract deterministic workbook artifacts from an Excel file into the planning workspace.

AICODE-NOTE: Workbook extraction stays deterministic and provenance-preserving so later signal detection can remain explainable and reviewer-friendly.
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter

from planning_workspace_lib import (
    ensure_attractor_stage_artifacts,
    load_run_manifest,
    relative_to_run,
    slugify,
    touch_generated_artifact,
    update_checkpoint,
    update_run_status,
    update_scenario_status,
    write_json,
    write_text,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract workbook artifacts into intake/.")
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--scenario-id", default="baseline")
    parser.add_argument("--stage-id", default="extract_workbook")
    return parser.parse_args()


def display_value(cell: Cell, value_cell: Cell | None = None) -> Any:
    value = value_cell.value if value_cell is not None else cell.value
    if value is None:
        return None
    if (value_cell and value_cell.is_date) or cell.is_date:
        if hasattr(value, "isoformat"):
            return value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def json_safe_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def inferred_type(cell: Cell) -> str:
    if cell.value is None:
        return "blank"
    if cell.data_type == "f":
        return "formula"
    if cell.is_date:
        return "date"
    if isinstance(cell.value, bool):
        return "boolean"
    if isinstance(cell.value, (int, float)):
        return "number"
    return "string"


def markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return "_No populated cells detected._"
    widths = [0] * len(rows[0])
    for row in rows:
        for index, value in enumerate(row):
            widths[index] = max(widths[index], len(value))
    header = "| " + " | ".join(value.ljust(widths[index]) for index, value in enumerate(rows[0])) + " |"
    separator = "| " + " | ".join("-" * widths[index] for index in range(len(rows[0]))) + " |"
    body = [
        "| " + " | ".join(value.ljust(widths[index]) for index, value in enumerate(row)) + " |"
        for row in rows[1:]
    ]
    return "\n".join([header, separator, *body])


def workbook_unsupported_features(source_path: Path) -> dict[str, bool]:
    with ZipFile(source_path) as archive:
        names = archive.namelist()
        return {
            "macros": source_path.suffix.lower() == ".xlsm" or any(name.endswith("vbaProject.bin") for name in names),
            "charts": any("/charts/" in name for name in names),
            "images": any("/media/" in name for name in names),
            "pivotTables": any("/pivotTables/" in name for name in names),
        }


def extract_sheet_rows(worksheet, value_worksheet) -> tuple[list[dict[str, Any]], list[list[str]]]:
    cell_records: list[dict[str, Any]] = []
    table_rows: list[list[str]] = []
    max_column = worksheet.max_column
    header = [get_column_letter(column) for column in range(1, max_column + 1)]
    table_rows.append(header)

    for row_index in range(1, worksheet.max_row + 1):
        row_values: list[str] = []
        for column_index in range(1, max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            value_cell = value_worksheet.cell(row=row_index, column=column_index)
            raw_value = cell.value
            shown = display_value(cell, value_cell)
            record = {
                "coordinate": cell.coordinate,
                "row": row_index,
                "column": column_index,
                "rawValue": json_safe_value(raw_value),
                "displayValue": shown,
                "formula": json_safe_value(raw_value) if cell.data_type == "f" else None,
                "inferredType": inferred_type(cell),
            }
            cell_records.append(record)
            row_values.append("" if shown is None else str(shown))
        if any(value != "" for value in row_values):
            table_rows.append(row_values)
    return cell_records, table_rows


def unique_sheet_slug(sheet_name: str, used_slugs: set[str]) -> str:
    base_slug = slugify(sheet_name)
    candidate = base_slug
    index = 2
    while candidate in used_slugs:
        candidate = f"{base_slug}-sheet-{index}"
        index += 1
    used_slugs.add(candidate)
    return candidate


def main() -> int:
    args = parse_args()
    run_dir = Path(args.run_dir).resolve()
    run_manifest = load_run_manifest(run_dir)
    primary_input = run_dir / run_manifest["primaryInputArtifact"]
    intake_root = run_dir / "intake"
    sheets_root = intake_root / "sheets"
    sheets_root.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(primary_input, data_only=False)
    workbook_values = load_workbook(primary_input, data_only=True)
    unsupported = workbook_unsupported_features(primary_input)
    workbook_manifest: dict[str, Any] = {
        "workbookFilename": primary_input.name,
        "sheetOrder": [],
        "hiddenSheetFlags": {},
        "mergedRanges": {},
        "formulaPresence": {},
        "dateLikeCells": {},
        "detectedTableRanges": {},
        "sheetArtifacts": {},
        "unsupportedFeatures": unsupported,
    }
    used_sheet_slugs: set[str] = set()

    for worksheet in workbook.worksheets:
        sheet_slug = unique_sheet_slug(worksheet.title, used_sheet_slugs)
        workbook_manifest["sheetOrder"].append(worksheet.title)
        workbook_manifest["hiddenSheetFlags"][worksheet.title] = worksheet.sheet_state != "visible"
        workbook_manifest["mergedRanges"][worksheet.title] = [str(item) for item in worksheet.merged_cells.ranges]
        workbook_manifest["formulaPresence"][worksheet.title] = any(
            cell.data_type == "f" for row in worksheet.iter_rows() for cell in row if cell.value is not None
        )
        workbook_manifest["dateLikeCells"][worksheet.title] = [
            cell.coordinate
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None and (cell.is_date or is_date_format(cell.number_format))
        ]
        workbook_manifest["detectedTableRanges"][worksheet.title] = [table.ref for table in worksheet.tables.values()]

        value_worksheet = workbook_values[worksheet.title]
        cell_records, table_rows = extract_sheet_rows(worksheet, value_worksheet)
        sheet_payload = {
            "sheetName": worksheet.title,
            "sheetSlug": sheet_slug,
            "hidden": worksheet.sheet_state != "visible",
            "rowBounds": {"min": worksheet.min_row, "max": worksheet.max_row},
            "columnBounds": {"min": worksheet.min_column, "max": worksheet.max_column},
            "mergedRanges": [str(item) for item in worksheet.merged_cells.ranges],
            "cells": cell_records,
        }
        json_path = sheets_root / f"{sheet_slug}.json"
        csv_path = sheets_root / f"{sheet_slug}.csv"
        md_path = sheets_root / f"{sheet_slug}.md"
        workbook_manifest["sheetArtifacts"][worksheet.title] = {
            "slug": sheet_slug,
            "jsonPath": f"intake/sheets/{sheet_slug}.json",
            "csvPath": f"intake/sheets/{sheet_slug}.csv",
            "markdownPath": f"intake/sheets/{sheet_slug}.md",
        }
        write_json(json_path, sheet_payload)
        with csv_path.open("w", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerows(table_rows)
        write_text(md_path, markdown_table(table_rows))
        for artifact in [json_path, csv_path, md_path]:
            touch_generated_artifact(run_dir, artifact)

    manifest_path = intake_root / "workbook-manifest.json"
    write_json(manifest_path, workbook_manifest)
    touch_generated_artifact(run_dir, manifest_path)

    update_run_status(
        run_dir,
        state="running",
        current_stage=args.stage_id,
        next_action="Build workbook inventory and region candidates",
        latest_summary=f"Extracted workbook intake for {primary_input.name}.",
        active_scenario_id=args.scenario_id,
    )
    update_checkpoint(
        run_dir,
        state="running",
        current_stage=args.stage_id,
        next_action="Build workbook inventory and region candidates",
        latest_summary=f"Extracted workbook intake for {primary_input.name}.",
        active_scenario_id=args.scenario_id,
        resume_hint="Run build_sheet_inventory next.",
    )
    update_scenario_status(
        run_dir,
        args.scenario_id,
        state="running",
        current_stage=args.stage_id,
        next_action="Build workbook inventory",
        latest_summary=f"Workbook extraction completed for {primary_input.name}.",
    )
    ensure_attractor_stage_artifacts(
        run_dir,
        stage_id=args.stage_id,
        command="extract_excel_artifacts.py",
        inputs={"runDir": str(run_dir), "workbook": str(primary_input)},
        summary="Workbook manifest and per-sheet JSON/CSV/Markdown artifacts generated.",
        state="success",
        outputs={"workbookManifest": relative_to_run(run_dir, manifest_path)},
    )

    print(f"WORKBOOK_MANIFEST={manifest_path}")
    print("STATE=running")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
