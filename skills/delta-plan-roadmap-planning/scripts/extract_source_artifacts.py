#!/usr/bin/env python3
"""Extract deterministic source artifacts into the planning workspace."""

from __future__ import annotations

import argparse
import csv
from datetime import datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter

from planning_workspace_lib import (
    load_run_manifest,
    relative_to_run,
    segment_artifacts_dir,
    slugify,
    source_manifest_path,
    touch_generated_artifact,
    update_checkpoint,
    update_run_status,
    update_scenario_status,
    write_json,
    write_text,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract source artifacts into intake/.")
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--scenario-id", default="baseline")
    parser.add_argument("--stage-id", default="extract_source_artifacts")
    return parser.parse_args()


def display_value(cell: Cell, value_cell: Cell | None = None) -> Any:
    value = value_cell.value if value_cell is not None else cell.value
    if value is None:
        return None
    if (value_cell and value_cell.is_date) or cell.is_date:
        if hasattr(value, "isoformat"):
            return value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def json_safe_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def inferred_type(cell: Cell) -> str:
    if cell.value is None:
        return "blank"
    if cell.data_type == "f":
        return "formula"
    if cell.is_date:
        return "date"
    if isinstance(cell.value, bool):
        return "boolean"
    if isinstance(cell.value, (int, float)):
        return "number"
    return "string"


def markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return "_No populated values detected._"
    widths = [0] * len(rows[0])
    for row in rows:
        for index, value in enumerate(row):
            widths[index] = max(widths[index], len(value))
    header = "| " + " | ".join(value.ljust(widths[index]) for index, value in enumerate(rows[0])) + " |"
    separator = "| " + " | ".join("-" * widths[index] for index in range(len(rows[0]))) + " |"
    body = [
        "| " + " | ".join(value.ljust(widths[index]) for index, value in enumerate(row)) + " |"
        for row in rows[1:]
    ]
    return "\n".join([header, separator, *body])


def excel_unsupported_features(source_path: Path) -> dict[str, bool]:
    with ZipFile(source_path) as archive:
        names = archive.namelist()
        return {
            "macros": source_path.suffix.lower() == ".xlsm" or any(name.endswith("vbaProject.bin") for name in names),
            "charts": any("/charts/" in name for name in names),
            "images": any("/media/" in name for name in names),
            "pivotTables": any("/pivotTables/" in name for name in names),
        }


def extract_tab_rows(worksheet, value_worksheet) -> tuple[list[dict[str, Any]], list[list[str]]]:
    cell_records: list[dict[str, Any]] = []
    table_rows: list[list[str]] = []
    max_column = worksheet.max_column
    header = [get_column_letter(column) for column in range(1, max_column + 1)]
    table_rows.append(header)

    for row_index in range(1, worksheet.max_row + 1):
        row_values: list[str] = []
        for column_index in range(1, max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            value_cell = value_worksheet.cell(row=row_index, column=column_index)
            raw_value = cell.value
            shown = display_value(cell, value_cell)
            record = {
                "coordinate": cell.coordinate,
                "row": row_index,
                "column": column_index,
                "rawValue": json_safe_value(raw_value),
                "displayValue": shown,
                "formula": json_safe_value(raw_value) if cell.data_type == "f" else None,
                "inferredType": inferred_type(cell),
            }
            cell_records.append(record)
            row_values.append("" if shown is None else str(shown))
        if any(value != "" for value in row_values):
            table_rows.append(row_values)
    return cell_records, table_rows


def unique_segment_slug(label: str, used_slugs: set[str]) -> str:
    base_slug = slugify(label)
    candidate = base_slug
    index = 2
    while candidate in used_slugs:
        candidate = f"{base_slug}-segment-{index}"
        index += 1
    used_slugs.add(candidate)
    return candidate


def write_segment_artifacts(run_dir: Path, segment_root: Path, segment_slug: str, segment_payload: dict[str, Any], table_rows: list[list[str]]) -> dict[str, str]:
    json_path = segment_root / f"{segment_slug}.json"
    csv_path = segment_root / f"{segment_slug}.csv"
    md_path = segment_root / f"{segment_slug}.md"
    write_json(json_path, segment_payload)
    with csv_path.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(table_rows)
    write_text(md_path, markdown_table(table_rows))
    for artifact in [json_path, csv_path, md_path]:
        touch_generated_artifact(run_dir, artifact)
    return {
        "jsonPath": relative_to_run(run_dir, json_path),
        "csvPath": relative_to_run(run_dir, csv_path),
        "markdownPath": relative_to_run(run_dir, md_path),
    }


def extract_excel_source(run_dir: Path, primary_input: Path, manifest: dict[str, Any]) -> None:
    segment_root = segment_artifacts_dir(run_dir)
    segment_root.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(primary_input, data_only=False)
    workbook_values = load_workbook(primary_input, data_only=True)
    used_segment_slugs: set[str] = set()

    manifest.update(
        {
            "sourceFilename": primary_input.name,
            "sourceKind": "excel_workbook",
            "containerOrder": [],
            "segmentArtifacts": {},
            "parserMetadata": {
                "tabOrder": [],
                "hiddenTabFlags": {},
                "mergedRanges": {},
                "formulaPresence": {},
                "dateLikeCells": {},
                "detectedTableRanges": {},
                "unsupportedFeatures": excel_unsupported_features(primary_input),
            },
        }
    )

    for worksheet in workbook.worksheets:
        segment_slug = unique_segment_slug(worksheet.title, used_segment_slugs)
        manifest["containerOrder"].append(worksheet.title)
        manifest["parserMetadata"]["tabOrder"].append(worksheet.title)
        manifest["parserMetadata"]["hiddenTabFlags"][worksheet.title] = worksheet.sheet_state != "visible"
        manifest["parserMetadata"]["mergedRanges"][worksheet.title] = [str(item) for item in worksheet.merged_cells.ranges]
        manifest["parserMetadata"]["formulaPresence"][worksheet.title] = any(
            cell.data_type == "f" for row in worksheet.iter_rows() for cell in row if cell.value is not None
        )
        manifest["parserMetadata"]["dateLikeCells"][worksheet.title] = [
            cell.coordinate
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None and (cell.is_date or is_date_format(cell.number_format))
        ]
        manifest["parserMetadata"]["detectedTableRanges"][worksheet.title] = [table.ref for table in worksheet.tables.values()]

        value_worksheet = workbook_values[worksheet.title]
        cell_records, table_rows = extract_tab_rows(worksheet, value_worksheet)
        segment_payload = {
            "segmentId": segment_slug,
            "containerLabel": worksheet.title,
            "segmentLabel": worksheet.title,
            "hidden": worksheet.sheet_state != "visible",
            "rowBounds": {"min": worksheet.min_row, "max": worksheet.max_row},
            "columnBounds": {"min": worksheet.min_column, "max": worksheet.max_column},
            "cells": cell_records,
            "parserMetadata": {
                "sourceKind": "excel_workbook",
                "tabName": worksheet.title,
                "mergedRanges": [str(item) for item in worksheet.merged_cells.ranges],
            },
        }
        manifest["segmentArtifacts"][segment_slug] = write_segment_artifacts(
            run_dir, segment_root, segment_slug, segment_payload, table_rows
        )


def extract_flat_table_source(run_dir: Path, primary_input: Path, manifest: dict[str, Any], source_kind: str) -> None:
    segment_root = segment_artifacts_dir(run_dir)
    segment_root.mkdir(parents=True, exist_ok=True)
    segment_slug = unique_segment_slug(primary_input.stem, set())
    rows = primary_input.read_text(encoding="utf-8", errors="ignore").splitlines()
    if source_kind == "csv":
        table_rows = list(csv.reader(rows))
    else:
        table_rows = [["lineNumber", "content"]] + [[str(index), line] for index, line in enumerate(rows, start=1)]

    segment_payload = {
        "segmentId": segment_slug,
        "containerLabel": primary_input.name,
        "segmentLabel": primary_input.stem,
        "hidden": False,
        "rowBounds": {"min": 1, "max": len(table_rows)},
        "columnBounds": {"min": 1, "max": max((len(row) for row in table_rows), default=1)},
        "rows": table_rows,
        "parserMetadata": {"sourceKind": source_kind},
    }
    manifest.update(
        {
            "sourceFilename": primary_input.name,
            "sourceKind": source_kind,
            "containerOrder": [primary_input.name],
            "segmentArtifacts": {
                segment_slug: write_segment_artifacts(run_dir, segment_root, segment_slug, segment_payload, table_rows)
            },
            "parserMetadata": {"sourceKind": source_kind},
        }
    )


def main() -> int:
    args = parse_args()
    run_dir = Path(args.run_dir).resolve()
    run_manifest = load_run_manifest(run_dir)
    primary_input = run_dir / run_manifest["primaryInputArtifact"]
    intake_manifest = source_manifest_path(run_dir)
    source_manifest = {
        "primarySourceId": run_manifest.get("primarySourceId"),
        "primaryInputArtifact": run_manifest.get("primaryInputArtifact"),
        "sourceKind": run_manifest.get("sourceKind"),
        "sources": run_manifest.get("sourceFiles", []),
        "segmentArtifacts": {},
    }

    source_kind = run_manifest.get("sourceKind")
    if source_kind == "excel_workbook":
        extract_excel_source(run_dir, primary_input, source_manifest)
    elif source_kind in {"csv", "markdown", "text"}:
        extract_flat_table_source(run_dir, primary_input, source_manifest, source_kind)
    else:
        raise SystemExit(f"Unsupported source kind for extraction: {source_kind}")

    write_json(intake_manifest, source_manifest)
    touch_generated_artifact(run_dir, intake_manifest)

    update_run_status(
        run_dir,
        state="running",
        current_stage=args.stage_id,
        next_action="Build source inventory and segment candidates",
        latest_summary=f"Extracted source intake for {primary_input.name}.",
        active_scenario_id=args.scenario_id,
    )
    update_checkpoint(
        run_dir,
        state="running",
        current_stage=args.stage_id,
        next_action="Build source inventory and segment candidates",
        latest_summary=f"Extracted source intake for {primary_input.name}.",
        active_scenario_id=args.scenario_id,
        resume_hint="Run build_source_inventory next.",
    )
    update_scenario_status(
        run_dir,
        args.scenario_id,
        state="running",
        current_stage=args.stage_id,
        next_action="Build source inventory",
        latest_summary=f"Source extraction completed for {primary_input.name}.",
    )

    print(f"SOURCE_MANIFEST={intake_manifest}")
    print("STATE=running")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
